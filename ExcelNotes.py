'''
Created on Feb 25, 2019

@author: darrenbean
'''
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

"""
print(type(wb))
print(wb.get_sheet_names()) # depreciated, wb.sheetnames
sheet = wb.get_sheet_by_name('Sheet3') # depreciated, wb[sheetname]
print(sheet)
print(type(sheet))
print(sheet.title)
anotherSheet = wb.active
print(anotherSheet)
"""

"""
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)
# note gotta use str converter if value is int
# site says col is letter val but py interprets it as column?
# can use get_column_letter(), column_index_from_string() to get (import 1st) 
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)

print(sheet.cell(row=1, column=2).value) # note numbering here starts at 1, not 0
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value) # gives all row i col 2 vals
"""

"""
print(sheet.max_row)
print(sheet.max_column)
print(get_column_letter(1))
print(column_index_from_string('AA'))
"""

"""
tuple(sheet['A1':'C3'])
# geneartes tuples for each row in slice
print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']: # goest through rows in slice
    for cellObj in rowOfCellObjects:   # goes through cell in row
        print(cellObj.coordinate, cellObj.value)  
    print('--- END OF ROW ---')
"""

#print(sheet.columns[1]) # this would be 2d column
#for cellObj in sheet.columns[1]:
#    print(cellObj.value)
# columns has changed, produces a generator so need to make a list or use next()

#this works
print(list(sheet.columns)[0]) #note since this is array index, starts at 0
for cellObj in list(sheet.columns)[0]:
    print(cellObj.value)

# both rows and cells gives tuple of tuples



    
    
