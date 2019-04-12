'''
Created on Mar 14, 2019

@author: darrenbean
'''

import openpyxl

"""
wb = openpyxl.Workbook()
sheet = wb.active
print(wb.get_sheet_names()) # wb.sheetnames
print(sheet.title)
sheet.title = 'spam Bacon Eggs Sheet'
print(wb.sheetnames)
"""

"""
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = "Spam Spam Spam"
wb.save('example_copy.xlsx') # giving different name saves changes to copy
"""

"""
wb = openpyxl.Workbook()
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index =0, title='First Sheet')
print(wb.sheetnames)
"""

"""
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Hello World!'
print(sheet['A1'].value)
"""

