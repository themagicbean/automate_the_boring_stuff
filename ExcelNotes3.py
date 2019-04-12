'''
Created on Mar 14, 2019

@author: darrenbean
'''

import openpyxl
from openpyxl.styles import Font # note capitalized
from openpyxl.chart import Reference, Series, BarChart

wb = openpyxl.Workbook()

"""
sheet = wb.get_sheet_by_name('Sheet') # wb.sheetname

fontObj1 = Font(name='TimesNewRoman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt italic'

wb.save('styles.xlsx')
"""

"""
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')
"""

"""
sheet = wb.get_sheet_by_name('Sheet') # adjusting rows and columns, about 3/4 down
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions[2] = 20
wb.template = False
wb.save('dimensions.xlsx')
wb.close()
# getting saving errors for some reason?
"""

"""
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = "Twelve cells merged together."
sheet.merge_cells('C5:D5')
sheet['C5'] = "Two merged cells."
wb.save('merged.xlsx')
# this one is fine though
"""

"""
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i

# get lots of red Xs for "chart" but it works
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)

seriesObj = openpyxl.chart.Series(refObj, title='First Series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')
"""






