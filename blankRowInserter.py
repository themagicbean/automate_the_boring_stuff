'''
Created on Mar 18, 2019

takes three argvs int1 int2, inserts int1 blank rows after int1 rows
third argv is name of file to have stuff inserted into
sargvs set << 8 3 produceSales.xlsx >> in run options in eclipse

@author: darrenbean
'''

import openpyxl
import sys


rowToSplit = int(sys.argv[1])
rowsToInsert = int(sys.argv[2])
fileToModify = sys.argv[3]

wb = openpyxl.load_workbook(fileToModify)
sheet = wb.active
sheet = sheet.insert_rows(rowToSplit+1, rowsToInsert)
wb.save("insertedRows.xlsx")
# works!


# all this was unnecssary b/c saving as eliminates need to copy.
"""
wbToModify = openpyxl.load_workbook(fileToModify)
sheetToModify = wbToModify.active

wbModified = openpyxl.Workbook()
sheetModified = wbModified.active
sheetModified = wbToModify.copy_worksheet(sheetToModify)
#sheetModified = sheetModified.insert_rows(rowToSplit+1, rowsToInsert)

wbModified.save('insertedRows.xlsx')
# not working gives a blank sheet error when copying
"""